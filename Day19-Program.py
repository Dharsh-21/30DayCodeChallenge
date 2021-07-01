#Day19
#1.	Create an Excel with data of Student database and add all the 
# values which is required for student management database, 
# Read the excel file and add the datas into a DB 
from openpyxl import Workbook

workbook = Workbook()
spreadsheet = workbook.active

spreadsheet["A1"] = "Name"
spreadsheet["B1"] = "Python"
spreadsheet["C1"] = "Chemistry"
spreadsheet["D1"] = "Tamil"
spreadsheet["A2"] = "vishika"
spreadsheet["B2"] = "90"
spreadsheet["C2"] = "100"
spreadsheet["D2"] = "98"
spreadsheet["A3"] = "Dharshka"
spreadsheet["B3"] = "96"
spreadsheet["C3"] = "90"
spreadsheet["D3"] = "95"

m_row = spreadsheet.max_row
# Loop will print all values
# of first column
for i in range(1, m_row + 1):
    cell_obj = spreadsheet.cell(row = i, column = 1)
    print(cell_obj.value)

workbook.save(filename="HelloWorld.xlsx")